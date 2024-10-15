'''
# -*- coding: UTF-8 -*-
Author: Yingyu Wang
LastEditTime: 2024-10-15 15:19:46
LastEditors: naivsheng naivsheng@outlook.com
Description: get当前文件夹内全部文件名以归档: 获取文件夹名并创建对应的sheet页, 将文件夹中的pdf文件汇总到表格中, 调整列宽
FilePath: \crawer\file_Velten.py
'''
def read_file():
    with open('files.txt','r',encoding='utf-8') as f:
        FLs = f.readlines()
    return FLs
def write_file(filiale):
    elements_with_char = [s for s in filiale if '.pdf' in s]
    for elem in elements_with_char:
        filiale.pop(filiale.index(elem))
    with open('files.txt','w',encoding='utf-8')as f:
        f.writelines(line + "\n" for line in filiale)
import os, datetime
import openpyxl
import openpyxl.worksheet
path = os.getcwd()
folder_list = []
record = {}
for file in os.listdir(path): # find all folders: folder_list = list()
    if os.path.isdir(file):
        folder_list.append(file)
wb = openpyxl.load_workbook('Velten.xlsx')
sheets_name = wb.sheetnames
with open('test.txt','w+',encoding='utf-8') as f:
    f.write(f'{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")} start writing\n')
for folder in folder_list:
    FLs = read_file()
    pdf_list = {x.replace('\n',''):[] for x in FLs} # get FL list
    if folder not in sheets_name:
        ws=wb.create_sheet(folder,0)
        headers = ['Filiale','Files']
        ws.append(headers)
    else:
        ws = wb[folder]
    target_path = f'{path}\\{folder}'
    output = os.listdir(target_path)
    menge = len(pdf_list)
    for file in output: # get pdf Datein
        if file.endswith('.pdf') or file.endswith('.PDF'):
            filiale = file.split('-')[0]
            if filiale == 'Stuttgart':
                filiale = 'Stuttgart-2'
            if 'Turmstr' in filiale:
                filiale = 'BerlinTurmstr'
            elif filiale not in pdf_list:
                pdf_list[filiale] = []
            pdf_list[filiale].append(file) 
    ws.delete_rows(2, ws.max_row - 1)
    with open('test.txt','a+',encoding='utf-8') as f:
        f.write(f'{folder}: \n')
    for fl,info in pdf_list.items(): # writedown
        row = [fl] + info
        with open('test.txt','a+',encoding='utf-8') as f:
            f.write(f'{row}\n')
        if info != []: # how many files for each Ordner is already update
            menge -= 1
        ws.append(row)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    record[folder] = menge
    print(record)
    write_file(list(pdf_list.keys()))
sheets = wb.sheetnames
if '目录' in sheets:
    del wb['目录']
    wb.save('Velten.xlsx') 
    sheets = wb.sheetnames
ws = wb.create_sheet('目录', 0)
for i, sheet_name in enumerate(sheets):
    if sheet_name == '目录':
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        continue
    elif sheet_name in folder_list:
        safe_sheet_name = f"'{sheet_name}'"
        # add hylink to other sheets
        cell = ws.cell(row=i + 1, column=1)
        cell.value = f"Link to {sheet_name}"  # set show Text
        cell.hyperlink = f"#{safe_sheet_name}!A1"  # set hylink
        cell.style = "Hyperlink"  
        cell = ws.cell(row=i + 1, column=2)
        cell.value = f"fehlt {record[sheet_name]} Filialen" 
        # add hylink to Menu
        worksheet = wb[sheet_name]
        return_cell = worksheet.cell(row=1, column=3)
        if return_cell.hyperlink is None or return_cell.hyperlink.target != "#'目录'!A1":
            return_cell.value = "返回目录"
            return_cell.hyperlink = "#'目录'!A1"  
            return_cell.style = "Hyperlink" 
# save
wb.save('Velten.xlsx')